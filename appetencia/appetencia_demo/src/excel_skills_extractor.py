# src/excel_skills_extractor.py
from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import openpyxl
from .utils import norm, fill_signature


@dataclass
class SkillEvidence:
    client_id: str
    activity_label: str
    sheet_name: str
    family: str
    skill: str
    dimension: str  # RESULTAT / MAITRISE / PLAISIR
    cell: str
    fill_sig: tuple
    source_file: str


LEGEND_KEYS = {
    "resultat": "RESULTAT",
    "résultat": "RESULTAT",
    "maitrisee": "MAITRISE",
    "maîtrisée": "MAITRISE",
    "plaisir": "PLAISIR",
}

# Liste des familles attendues dans tes tableaux (à ajuster si besoin)
FAMILIES = [
    "Contrôler",
    "Développer",
    "Conseiller",
    "Créer, produire",
    "Organiser, gérer",
    "Chercher",
    "Communiquer",
    "Décider",
]

FAMILIES_NORM = {norm(x): x for x in FAMILIES}


def read_activity_title(ws) -> Optional[str]:
    for coord in ("C1", "B1", "A1"):
        v = ws[coord].value
        if isinstance(v, str) and v.strip():
            return v.strip()
    return None


def read_legend_fill(ws) -> Dict[tuple, str]:
    """
    Retourne mapping: fill_signature -> dimension
    en cherchant les cellules 'Résultat', 'Maîtrisée', 'Plaisir'
    """
    sig_to_dim: Dict[tuple, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            key = norm(cell.value)
            if key in LEGEND_KEYS:
                sig = fill_signature(cell)
                if sig:
                    sig_to_dim[sig] = LEGEND_KEYS[key]
    return sig_to_dim


def find_family_header_row(ws, max_rows: int = 60) -> Tuple[Optional[int], Dict[int, str]]:
    """
    Trouve la ligne qui contient plusieurs familles connues.
    Retourne (row_index, {col_index: family})
    """
    best_row = None
    best_map: Dict[int, str] = {}

    for r in range(1, min(ws.max_row, max_rows) + 1):
        colmap: Dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            vn = norm(v)
            if vn in FAMILIES_NORM:
                colmap[c] = FAMILIES_NORM[vn]

        # on veut une ligne qui en contient plusieurs
        if len(colmap) >= 3:
            if best_row is None or len(colmap) > len(best_map):
                best_row = r
                best_map = colmap

    return best_row, best_map


def extract_skills_from_workbook(filepath: str, client_id: str) -> List[SkillEvidence]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_evidence: List[SkillEvidence] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if norm(sheet_name) in {"qualites", "qualités"}:
            continue

        sig_to_dim = read_legend_fill(ws)
        header_row, col_to_family = find_family_header_row(ws)

        # DEBUG utile si encore 0 (tu peux laisser)
        print(sheet_name, "legend:", len(sig_to_dim), "headers:", len(col_to_family), "row:", header_row)

        if not sig_to_dim or not header_row or not col_to_family:
            continue

        activity_label = read_activity_title(ws) or sheet_name

        for r in range(header_row + 1, ws.max_row + 1):
            for c, family in col_to_family.items():
                cell = ws.cell(r, c)
                if not isinstance(cell.value, str):
                    continue
                skill = cell.value.strip()
                if not skill:
                    continue

                sig = fill_signature(cell)
                if not sig:
                    continue

                dim = sig_to_dim.get(sig)
                if not dim:
                    continue

                all_evidence.append(
                    SkillEvidence(
                        client_id=client_id,
                        activity_label=activity_label,
                        sheet_name=sheet_name,
                        family=family,
                        skill=skill,
                        dimension=dim,
                        cell=cell.coordinate,
                        fill_sig=sig,
                        source_file=filepath,
                    )
                )

    return all_evidence
