# src/finalites_extractor.py
from __future__ import annotations
from dataclasses import dataclass
from typing import List, Optional
import os
import openpyxl

from .utils import norm


@dataclass
class FinaliteSelection:
    client_id: str
    categorie: str
    finalite: str
    source_file: str
    sheet: str


def _is_category_row(text: str) -> bool:
    """
    Catégories typiques : tout en majuscules, ex 'LE RELATIONNEL ET L’HUMAIN'
    """
    t = text.strip()
    return len(t) >= 8 and t.upper() == t


def extract_finalites_from_file(filepath: str, client_id: str) -> List[FinaliteSelection]:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    results: List[FinaliteSelection] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        categorie = "NON_CLASSEE"

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
            a = row[0].value  # colonne A: texte
            c = row[2].value  # colonne C: coche souvent 'x'

            if isinstance(a, str) and a.strip():
                # mise à jour catégorie
                if _is_category_row(a):
                    categorie = a.strip()
                    continue

                # sélection = si colonne C contient 'x' (insensible à la casse)
                selected = isinstance(c, str) and norm(c) in {"x", "oui", "ok"}

                if selected:
                    results.append(
                        FinaliteSelection(
                            client_id=client_id,
                            categorie=categorie,
                            finalite=a.strip(),
                            source_file=filepath,
                            sheet=sheet_name,
                        )
                    )

    return results
