from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, Optional, Set, Tuple
import re
import openpyxl

from .utils import norm


def _aliases_from_motor(motor_short: str) -> Set[str]:
    """
    Alias à partir du verbe moteur court.
    ex: "Rechercher" -> {"rechercher"}
    """
    t = norm(motor_short)
    aliases = set()
    if t:
        aliases.add(t)
        aliases.add(t.replace(" ", ""))
    return aliases


@dataclass(frozen=True)
class FinalitesRepo:
    alias_to_finalite: Dict[str, str]    # alias_norm -> finalité
    finalites: Set[str]


def load_finalites_repo(path: str) -> FinalitesRepo:
    """
    Charge un fichier Excel 'finalités motrices'.
    Supporte 2 formats fréquents :

    FORMAT A (table) :
      colonne A = verbe moteur (ou catégorie)
      colonne B = finalité
    ou l’inverse.

    FORMAT B (par colonnes) :
      header de colonne = finalité
      cellules dessous = verbes moteurs associés.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    alias_to_finalite: Dict[str, str] = {}
    finalites: Set[str] = set()

    for sname in wb.sheetnames:
        ws = wb[sname]

        # On lit toutes les cellules en mémoire (valeurs uniquement)
        grid = [[cell.value for cell in row] for row in ws.iter_rows()]

        # --------- Tentative FORMAT B (header = finalité) ----------
        # On regarde la première ligne : si elle contient plusieurs textes, on l’interprète comme headers
        if grid and any(isinstance(v, str) and v.strip() for v in grid[0]):
            headers = [v.strip() if isinstance(v, str) else None for v in grid[0]]
            # Si on a au moins 2 headers texte, on tente le format par colonnes
            nb_headers_text = sum(1 for h in headers if isinstance(h, str) and h.strip())
            if nb_headers_text >= 2:
                for col_idx, header in enumerate(headers):
                    if not isinstance(header, str) or not header.strip():
                        continue
                    finalite = header.strip()
                    finalites.add(finalite)
                    # cellules sous le header = verbes moteurs
                    for r in range(1, len(grid)):
                        v = grid[r][col_idx] if col_idx < len(grid[r]) else None
                        if isinstance(v, str) and v.strip():
                            motor = v.strip()
                            for al in _aliases_from_motor(motor):
                                alias_to_finalite.setdefault(al, finalite)

        # --------- Tentative FORMAT A (table 2 colonnes) ----------
        # On parcourt les lignes et on cherche des paires texte/texte
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            texts = [v.strip() for v in row if isinstance(v, str) and v.strip()]
            if len(texts) < 2:
                continue

            # Heuristique : prendre les deux premiers textes de la ligne comme paire
            a, b = texts[0], texts[1]

            # On ne sait pas si c'est (motor -> finalite) ou (finalite -> motor)
            # On décide en utilisant une heuristique : une finalité est souvent "nom abstrait" (plusieurs mots),
            # un verbe moteur est souvent 1 mot capitalisé.
            a_is_short = len(a.split()) <= 2
            b_is_short = len(b.split()) <= 2

            if a_is_short and not b_is_short:
                motor, finalite = a, b
            elif b_is_short and not a_is_short:
                motor, finalite = b, a
            else:
                # fallback : considérer A comme motor, B comme finalité
                motor, finalite = a, b

            finalites.add(finalite)
            for al in _aliases_from_motor(motor):
                alias_to_finalite.setdefault(al, finalite)

    return FinalitesRepo(alias_to_finalite=alias_to_finalite, finalites=finalites)


def map_motor_to_finalite(motor_short: str, repo: FinalitesRepo) -> str:
    """
    Mapping finalité depuis un verbe moteur court.
    """
    if not motor_short:
        return "Inconnue"

    key = norm(motor_short)
    if key in repo.alias_to_finalite:
        return repo.alias_to_finalite[key]

    key2 = key.replace(" ", "")
    if key2 in repo.alias_to_finalite:
        return repo.alias_to_finalite[key2]

    return "Inconnue"
