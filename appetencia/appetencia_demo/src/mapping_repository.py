# src/mapping_repository.py
from __future__ import annotations
from dataclasses import dataclass
from typing import Dict, Set, Optional
import re
import openpyxl
from .utils import norm


def _split_aliases(text: str) -> Set[str]:
    """
    Extrait des alias depuis un libellé de verbe moteur.
    Ex: "Exécuter : mettre en place..." -> {"executer", "mettre", "place", ...}
    Ex: "Améliorer / Optimiser" -> {"ameliorer", "optimiser"}
    """
    t = norm(text)
    # on coupe sur : / , ; ( ) - etc
    parts = re.split(r"[:/;,()\-\u2013\u2014]+", t)
    aliases: Set[str] = set()
    for p in parts:
        for w in p.split():
            if len(w) >= 4:  # évite "de", "la", etc.
                aliases.add(w)
    # on garde aussi la première “expression” si courte
    if parts and parts[0].strip():
        first = parts[0].strip()
        if len(first.split()) <= 3:
            aliases.add(first.replace(" ", ""))
            aliases.add(first.split()[0])
    return aliases


@dataclass(frozen=True)
class MotorRepo:
    motor_verbs: Set[str]                 # libellés originaux
    alias_to_motor: Dict[str, str]        # alias_norm -> libellé canonique
    family_fallback: Dict[str, str]       # famille_norm -> meilleur verbe moteur


def load_motor_repo(path: str) -> MotorRepo:
    wb = openpyxl.load_workbook(path, data_only=True)

    motor_verbs: Set[str] = set()
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.strip():
                    motor_verbs.add(v.strip())

    # index alias -> verbe moteur
    alias_to_motor: Dict[str, str] = {}
    for mv in motor_verbs:
        for al in _split_aliases(mv):
            # on garde le 1er rencontré si collision (OK pour une V2 démo)
            alias_to_motor.setdefault(al, mv)

    # fallback par famille : on choisit un verbe moteur “proche” via mots-clés
    # (mots-clés simples, on peut enrichir ensuite)
    family_keywords = {
        "communiquer": ["communi", "echang", "discut", "redig", "present", "sensibil", "faireconn"],
        "organiser, gerer": ["organis", "coordonn", "planif", "prevoir", "gerer"],
        "creer, produire": ["realiser", "fabriquer", "imaginer", "mettreforme", "produire", "elabor"],
        "developper": ["amelior", "optimis", "develop", "progres", "adapter"],
        "controler": ["control", "verif", "valider", "tester", "supervis", "diagnostiq"],
        "chercher": ["recherch", "etudier", "explor", "prospect", "investig", "interrog", "simuler"],
        "conseiller": ["conseill", "recommand", "preconis", "donneravis", "accompagn"],
        "decider": ["tranch", "decid", "choisir", "eliminer", "arbitr"],
    }

    def best_for_family(fam: str) -> Optional[str]:
        fn = norm(fam)
        # on cherche un verbe moteur qui contient des alias liés à la famille
        keys = family_keywords.get(fn, [])
        # score simple: nombre de mots-clés retrouvés dans les alias du verbe moteur
        best_mv, best_score = None, 0
        for mv in motor_verbs:
            mvn = norm(mv).replace(" ", "")
            score = sum(1 for k in keys if k in mvn)
            if score > best_score:
                best_score = score
                best_mv = mv
        return best_mv

    family_fallback: Dict[str, str] = {}
    for fam in family_keywords.keys():
        mv = best_for_family(fam)
        if mv:
            family_fallback[fam] = mv

    return MotorRepo(
        motor_verbs=motor_verbs,
        alias_to_motor=alias_to_motor,
        family_fallback=family_fallback,
    )


def map_skill_to_motor(skill: str, family: str, repo: MotorRepo) -> str:
    """
    V2 mapping :
    1) match exact sur base (1er mot) via alias_to_motor
    2) match sur mot-clé contenu (ex: "donner un avis")
    3) fallback par famille
    """
    if not skill:
        return "Inconnu"

    base = norm(skill.strip().split()[0])
    # match direct
    if base in repo.alias_to_motor:
        return repo.alias_to_motor[base]

    # match sur phrase complète normalisée (utile pour "donner un avis")
    full = norm(skill).replace(" ", "")
    if full in repo.alias_to_motor:
        return repo.alias_to_motor[full]

    # match par inclusion (très utile)
    for al, mv in repo.alias_to_motor.items():
        if al and al in full and len(al) >= 5:
            return mv

    # fallback par famille
    fam = norm(family)
    if fam in repo.family_fallback:
        return repo.family_fallback[fam]

    return "Inconnu"
