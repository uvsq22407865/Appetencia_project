import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ---------- Helpers ----------
def slug(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def is_anecdote_sheet(name: str) -> bool:
    try:
        int(str(name).strip())
        return True
    except ValueError:
        return False


# ---------- Files (same folder as this script) ----------
COMP_XLSX = Path("Tableau de compétences - LS -positionnement.xlsx")
FIN_XLSX = Path("MP_Liste des finalités motrices.xlsx")

OUT_DIR = Path("neo4j_csv_out")
OUT_DIR.mkdir(exist_ok=True)

CANDIDAT_ID = "C001"
CANDIDAT_NOM = "Candidat 1"

# ---------- Color mapping (YOUR rule) ----------
# Rouge = plaisir, Vert = resultat, Bleu = maitrise
COLOR_TO_GROUPE = {
    "FFFF0000": "PLAISIR",
    "FF92D050": "RESULTAT",
    "FF00B0F0": "MAITRISE",
}

# ---------- Output containers ----------
candidats = [{"candidatId": CANDIDAT_ID, "nom": CANDIDAT_NOM}]
anecdotes = []

categories = {}      # catId -> catName
competences = {}     # compId -> (compName, catId)
anecdote_comp = []   # rows: anecdoteId, competenceId, groupe, sourceCell

themes = {}          # themeId -> themeName
finalites = {}       # finId -> (finName, themeId)
priorites = []       # candidatId, finaliteId, niveau


# =========================
# 1) Competences (openpyxl: reads colors)
# =========================
if not COMP_XLSX.exists():
    raise FileNotFoundError(f"Fichier introuvable: {COMP_XLSX.resolve()}")

wb = load_workbook(COMP_XLSX)

for sh in wb.sheetnames:
    if not is_anecdote_sheet(sh):
        continue

    ws = wb[sh]
    sheet_num = str(sh).strip()
    anecdote_id = f"{CANDIDAT_ID}_SHEET_{sheet_num}"

    # Title is in C1 (row=1 col=3)
    titre_cell = ws.cell(row=1, column=3).value
    titre = str(titre_cell).strip() if titre_cell else f"Anecdote {sheet_num}"

    anecdotes.append(
        {
            "anecdoteId": anecdote_id,
            "candidatId": CANDIDAT_ID,
            "sheet": sheet_num,
            "titre": titre,
            "sourceFile": COMP_XLSX.name,
        }
    )

    # ---- Detect category columns from row 3, starting at column C (3)
    CAT_ROW = 3
    FIRST_COL = 3

    col_to_cat = {}
    empty_streak = 0
    for col in range(FIRST_COL, 80):  # wide enough
        val = ws.cell(row=CAT_ROW, column=col).value
        if val is None or str(val).strip() == "":
            empty_streak += 1
            if empty_streak >= 8:
                break
            continue

        empty_streak = 0
        cat_name = str(val).strip()
        cat_id = slug(cat_name)
        categories[cat_id] = cat_name
        col_to_cat[col] = cat_id

    if not col_to_cat:
        continue

    # ---- Scan competence cells under categories
    # Keep ONLY colored cells whose RGB matches COLOR_TO_GROUPE
    empty_row_streak = 0
    for r in range(CAT_ROW + 1, CAT_ROW + 250):
        any_value = False

        for col, cat_id in col_to_cat.items():
            cell = ws.cell(row=r, column=col)
            if cell.value is None or str(cell.value).strip() == "":
                continue

            any_value = True

            fill = cell.fill
            if not fill or fill.patternType is None:
                continue
            if str(fill.patternType).lower() != "solid":
                continue

            rgb = getattr(fill.fgColor, "rgb", None)
            if not isinstance(rgb, str):
                continue
            if rgb not in COLOR_TO_GROUPE:
                continue

            groupe = COLOR_TO_GROUPE[rgb]
            comp_name = str(cell.value).strip()
            comp_id = slug(comp_name)

            if comp_id not in competences:
                competences[comp_id] = (comp_name, cat_id)

            anecdote_comp.append(
                {
                    "anecdoteId": anecdote_id,
                    "competenceId": comp_id,
                    "groupe": groupe,
                    "sourceCell": f"{sheet_num}:R{r}C{col}",
                }
            )

        if not any_value:
            empty_row_streak += 1
            if empty_row_streak >= 25:
                break
        else:
            empty_row_streak = 0


# =========================
# 2) Finalites (pandas)
# =========================
if not FIN_XLSX.exists():
    raise FileNotFoundError(f"Fichier introuvable: {FIN_XLSX.resolve()}")

df_fin = pd.read_excel(FIN_XLSX, sheet_name="Finalités", header=None)

current_theme_id = None
for i in range(df_fin.shape[0]):
    c0 = df_fin.iloc[i, 0] if df_fin.shape[1] > 0 else None
    c1 = df_fin.iloc[i, 1] if df_fin.shape[1] > 1 else None
    c2 = df_fin.iloc[i, 2] if df_fin.shape[1] > 2 else None

    # Theme header row
    if pd.notna(c0) and str(c1).strip() == "Attention" and str(c2).strip() == "Important":
        theme_name = str(c0).strip()
        current_theme_id = slug(theme_name)
        themes[current_theme_id] = theme_name
        continue

    if pd.isna(c0) or current_theme_id is None:
        continue

    fin_name = str(c0).strip()
    if not fin_name:
        continue

    fin_id = "FIN_" + slug(fin_name)[:40]
    if fin_id not in finalites:
        finalites[fin_id] = (fin_name, current_theme_id)

    att = str(c1).strip().lower() == "x" if pd.notna(c1) else False
    imp = str(c2).strip().lower() == "x" if pd.notna(c2) else False

    if imp:
        priorites.append({"candidatId": CANDIDAT_ID, "finaliteId": fin_id, "niveau": "IMPORTANT"})
    if att:
        priorites.append({"candidatId": CANDIDAT_ID, "finaliteId": fin_id, "niveau": "ATTENTION"})


# =========================
# Write CSVs
# =========================
pd.DataFrame(candidats).to_csv(OUT_DIR / "candidats.csv", index=False)
pd.DataFrame(anecdotes).to_csv(OUT_DIR / "anecdotes.csv", index=False)

pd.DataFrame([{"categorieId": k, "nom": v} for k, v in categories.items()]).to_csv(
    OUT_DIR / "categories.csv", index=False
)

pd.DataFrame(
    [{"competenceId": k, "nom": v[0], "categorieId": v[1]} for k, v in competences.items()]
).to_csv(OUT_DIR / "competences.csv", index=False)

pd.DataFrame(anecdote_comp).to_csv(OUT_DIR / "anecdote_competence.csv", index=False)

pd.DataFrame([{"themeId": k, "nom": v} for k, v in themes.items()]).to_csv(
    OUT_DIR / "themes_finalites.csv", index=False
)

pd.DataFrame(
    [{"finaliteId": k, "nom": v[0], "themeId": v[1]} for k, v in finalites.items()]
).to_csv(OUT_DIR / "finalites.csv", index=False)

pd.DataFrame(priorites).to_csv(OUT_DIR / "priorites_finalites.csv", index=False)

print("OK ✅ CSV exportés dans:", OUT_DIR.resolve())
print("NB relations MONTRE (cellules colorées uniquement):", len(anecdote_comp))
print("NB compétences uniques:", len(competences))
