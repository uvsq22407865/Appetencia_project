import json
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# ---------- Helpers ----------
def slug(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r"[^A-Z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def norm_noaccent_lower(s: str) -> str:
    s = "" if s is None else str(s)
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


def is_anecdote_sheet(name: str) -> bool:
    try:
        int(str(name).strip())
        return True
    except ValueError:
        return False


def guess_groupe_from_text(v: str) -> str | None:
    """Accepte 'Résultat', 'resultats', 'MAITRISE', 'plaisir', etc."""
    t = norm_noaccent_lower(v)
    if "result" in t:
        return "RESULTAT"
    if "maitr" in t:
        return "MAITRISE"
    if "plais" in t:
        return "PLAISIR"
    return None


def extract_name_before_x(line: str) -> str:
    """
    Exemples:
      'Anticiper x 5 + 1' -> 'Anticiper'
      'Présenter x2' -> 'Présenter' (si séparateur pas parfait)
    """
    line = line.strip()
    if not line:
        return ""
    # split tolerant on " x " / "x" / "X"
    m = re.split(r"\s*[xX]\s*\d+.*$", line, maxsplit=1)
    if m and m[0].strip():
        return m[0].strip()
    # fallback: split on " x " even without digits
    if " x " in line.lower():
        return line.split(" x ")[0].strip()
    return line.strip()


# ---------- Color mapping (YOUR rule) ----------
# Rouge = plaisir, Vert = resultat, Bleu = maitrise
COLOR_TO_GROUPE = {
    "FFFF0000": "PLAISIR",
    "FF92D050": "RESULTAT",
    "FF00B0F0": "MAITRISE",
}

# ---------- Input / Output ----------
BASE_DIR = Path("data_candidats")  # dossier contenant C001/, C002/, ...
OUT_DIR = Path("neo4j_csv_out")
OUT_DIR.mkdir(exist_ok=True)

# ---------- Global containers (tous candidats) ----------
candidats = []
anecdotes = []

categories: dict[str, str] = {}          # catId -> catName (global)
competences: dict[str, tuple[str, str]] = {}  # compId -> (compName, catId) (global)
anecdote_comp = []                        # rows global

themes: dict[str, str] = {}              # themeId -> themeName (global)
finalites: dict[str, tuple[str, str]] = {}    # finId -> (finName, themeId) (global)
priorites = []                            # rows global


def load_candidate_meta(cand_dir: Path, cand_id: str) -> dict:
    """Optionnel: lire meta.json pour le nom du candidat."""
    meta_file = cand_dir / "meta.json"
    if meta_file.exists():
        try:
            return json.loads(meta_file.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {"nom": cand_id}


def process_competences_format_a(cand_id: str, comp_xlsx: Path, wb, anecdote_sheets: list[str]):
    """Format A: sheets numériques, compétences = cellules colorées."""
    for sh in anecdote_sheets:
        ws = wb[sh]
        sheet_num = str(sh).strip()
        anecdote_id = f"{cand_id}_SHEET_{sheet_num}"

        titre_cell = ws.cell(row=1, column=3).value  # C1
        titre = str(titre_cell).strip() if titre_cell else f"Anecdote {sheet_num}"

        anecdotes.append({
            "anecdoteId": anecdote_id,
            "candidatId": cand_id,
            "sheet": sheet_num,
            "titre": titre,
            "sourceFile": comp_xlsx.name,
        })

        # catégories ligne 3 (C..)
        CAT_ROW = 3
        FIRST_COL = 3

        col_to_cat = {}
        empty_streak = 0
        for col in range(FIRST_COL, 80):
            val = ws.cell(row=CAT_ROW, column=col).value
            if val is None or str(val).strip() == "":
                empty_streak += 1
                if empty_streak >= 8:
                    break
                continue

            empty_streak = 0
            cat_name = str(val).strip()
            cat_id = slug(cat_name)
            categories[cat_id] = cat_name
            col_to_cat[col] = cat_id

        if not col_to_cat:
            continue

        # scan compétences colorées
        empty_row_streak = 0
        for r in range(CAT_ROW + 1, CAT_ROW + 250):
            any_value = False
            for col, cat_id in col_to_cat.items():
                cell = ws.cell(row=r, column=col)
                if cell.value is None or str(cell.value).strip() == "":
                    continue

                any_value = True

                fill = cell.fill
                if not fill or fill.patternType is None:
                    continue
                if str(fill.patternType).lower() != "solid":
                    continue

                rgb = getattr(fill.fgColor, "rgb", None)
                if not isinstance(rgb, str) or rgb not in COLOR_TO_GROUPE:
                    continue

                groupe = COLOR_TO_GROUPE[rgb]
                comp_name = str(cell.value).strip()
                comp_id = slug(comp_name)

                if comp_id not in competences:
                    competences[comp_id] = (comp_name, cat_id)

                anecdote_comp.append({
                    "anecdoteId": anecdote_id,
                    "competenceId": comp_id,
                    "groupe": groupe,
                    "sourceCell": f"{cand_id}:{sheet_num}:R{r}C{col}",
                })

            if not any_value:
                empty_row_streak += 1
                if empty_row_streak >= 25:
                    break
            else:
                empty_row_streak = 0


def find_recap_sheet(wb) -> str | None:
    """
    Trouve un onglet qui ressemble à 'récap compétences' (robuste accents/majuscules).
    Si rien de clair, retourne None.
    """
    best = None
    for sh in wb.sheetnames:
        s = norm_noaccent_lower(sh)
        if "recap" in s and ("competence" in s or "competences" in s):
            return sh
        # fallback: parfois juste "competences"
        if best is None and ("competence" in s or "competences" in s):
            best = sh
    return best


def process_competences_format_b(cand_id: str, comp_xlsx: Path, wb):
    """
    Format B: pas de sheets numériques.
    On lit un onglet récap et on crée une pseudo-anecdote Cxxx_RECAP.
    Hypothèses (génériques) :
      - ligne 1 contient les catégories à partir de col B
      - col A contient le groupe (Plaisir/Maitrise/Resultat) ou une variante
      - les cellules contiennent des lignes type 'Anticiper x 5 + 1'
    """
    recap = find_recap_sheet(wb)
    if recap is None:
        print(f"[{cand_id}] ⚠️ Aucun sheet numérique et aucun recap compétences détecté dans {comp_xlsx.name} -> compétences non extraites")
        return

    ws = wb[recap]
    anecdote_id = f"{cand_id}_RECAP"

    anecdotes.append({
        "anecdoteId": anecdote_id,
        "candidatId": cand_id,
        "sheet": recap,
        "titre": "Récap compétences",
        "sourceFile": comp_xlsx.name,
    })

    # catégories sur la première ligne, à partir de B
    header_row = 1
    for col in range(2, 200):
        cat_val = ws.cell(row=header_row, column=col).value
        if cat_val is None or str(cat_val).strip() == "":
            continue

        cat_name = str(cat_val).strip()
        cat_id = slug(cat_name)
        categories[cat_id] = cat_name

        # lignes: col A = groupe ; col 'col' = texte compétences
        for r in range(2, 300):
            g_val = ws.cell(row=r, column=1).value
            if g_val is None or str(g_val).strip() == "":
                continue

            groupe = guess_groupe_from_text(str(g_val))
            if groupe is None:
                continue

            txt = ws.cell(row=r, column=col).value
            if txt is None or str(txt).strip() == "":
                continue

            # plusieurs items possibles dans la cellule
            for line in str(txt).splitlines():
                name = extract_name_before_x(line)
                if not name:
                    continue

                comp_name = name
                comp_id = slug(comp_name)

                if comp_id not in competences:
                    competences[comp_id] = (comp_name, cat_id)

                anecdote_comp.append({
                    "anecdoteId": anecdote_id,
                    "competenceId": comp_id,
                    "groupe": groupe,
                    "sourceCell": f"{cand_id}:{recap}:R{r}C{col}",
                })


def process_competences(cand_id: str, comp_xlsx: Path):
    if not comp_xlsx.exists():
        raise FileNotFoundError(f"[{cand_id}] Fichier compétences introuvable: {comp_xlsx.resolve()}")

    wb = load_workbook(comp_xlsx)
    anecdote_sheets = [sh for sh in wb.sheetnames if is_anecdote_sheet(sh)]

    if anecdote_sheets:
        process_competences_format_a(cand_id, comp_xlsx, wb, anecdote_sheets)
    else:
        process_competences_format_b(cand_id, comp_xlsx, wb)


def process_finalites(cand_id: str, fin_xlsx: Path):
    if not fin_xlsx.exists():
        raise FileNotFoundError(f"[{cand_id}] Fichier finalités introuvable: {fin_xlsx.resolve()}")

    xl = pd.ExcelFile(fin_xlsx)

    # Cherche un sheet qui contient "finalit" (avec ou sans accent)
    target = None
    for sh in xl.sheet_names:
        sh_norm = norm_noaccent_lower(sh)
        if "finalit" in sh_norm:
            target = sh
            break

    # Si rien trouvé, on prend le premier sheet
    if target is None:
        target = xl.sheet_names[0]

    df_fin = pd.read_excel(fin_xlsx, sheet_name=target, header=None)
    print(f"[{cand_id}] Finalites: lecture du sheet '{target}' dans {fin_xlsx.name}")

    current_theme_id = None
    for i in range(df_fin.shape[0]):
        c0 = df_fin.iloc[i, 0] if df_fin.shape[1] > 0 else None
        c1 = df_fin.iloc[i, 1] if df_fin.shape[1] > 1 else None
        c2 = df_fin.iloc[i, 2] if df_fin.shape[1] > 2 else None

        # entête thème (colonnes B/C = Attention/Important)
        if pd.notna(c0) and str(c1).strip() == "Attention" and str(c2).strip() == "Important":
            theme_name = str(c0).strip()
            current_theme_id = slug(theme_name)
            themes[current_theme_id] = theme_name
            continue

        if pd.isna(c0) or current_theme_id is None:
            continue

        fin_name = str(c0).strip()
        if not fin_name:
            continue

        fin_id = "FIN_" + slug(fin_name)[:40]
        if fin_id not in finalites:
            finalites[fin_id] = (fin_name, current_theme_id)

        att = str(c1).strip().lower() == "x" if pd.notna(c1) else False
        imp = str(c2).strip().lower() == "x" if pd.notna(c2) else False

        if imp:
            priorites.append({"candidatId": cand_id, "finaliteId": fin_id, "niveau": "IMPORTANT"})
        if att:
            priorites.append({"candidatId": cand_id, "finaliteId": fin_id, "niveau": "ATTENTION"})


# =========================
# Main: loop candidates
# =========================
if not BASE_DIR.exists():
    raise FileNotFoundError(f"Dossier introuvable: {BASE_DIR.resolve()}")

cand_dirs = sorted([p for p in BASE_DIR.iterdir() if p.is_dir() and p.name.startswith("C")])
if not cand_dirs:
    raise FileNotFoundError(f"Aucun dossier candidat trouvé dans {BASE_DIR}. Ex: data_candidats/C001/")

for cand_dir in cand_dirs:
    cand_id = cand_dir.name  # ex: C001
    meta = load_candidate_meta(cand_dir, cand_id)
    cand_name = meta.get("nom", cand_id)

    comp_xlsx = cand_dir / "competences.xlsx"
    fin_xlsx = cand_dir / "finalites.xlsx"

    candidats.append({"candidatId": cand_id, "nom": cand_name})

    process_competences(cand_id, comp_xlsx)
    process_finalites(cand_id, fin_xlsx)

print("Candidats traités:", len(candidats))

# =========================
# Write CSVs (global)
# =========================
pd.DataFrame(candidats).to_csv(OUT_DIR / "candidats.csv", index=False)
pd.DataFrame(anecdotes).to_csv(OUT_DIR / "anecdotes.csv", index=False)

pd.DataFrame([{"categorieId": k, "nom": v} for k, v in categories.items()]).to_csv(
    OUT_DIR / "categories.csv", index=False
)

pd.DataFrame(
    [{"competenceId": k, "nom": v[0], "categorieId": v[1]} for k, v in competences.items()]
).to_csv(OUT_DIR / "competences.csv", index=False)

pd.DataFrame(anecdote_comp).to_csv(OUT_DIR / "anecdote_competence.csv", index=False)

pd.DataFrame([{"themeId": k, "nom": v} for k, v in themes.items()]).to_csv(
    OUT_DIR / "themes_finalites.csv", index=False
)

pd.DataFrame(
    [{"finaliteId": k, "nom": v[0], "themeId": v[1]} for k, v in finalites.items()]
).to_csv(OUT_DIR / "finalites.csv", index=False)

pd.DataFrame(priorites).to_csv(OUT_DIR / "priorites_finalites.csv", index=False)

print("OK ✅ CSV exportés dans:", OUT_DIR.resolve())
print("NB anecdotes:", len(anecdotes))
print("NB relations MONTRE:", len(anecdote_comp))
print("NB compétences uniques:", len(competences))
print("NB priorités:", len(priorites))
